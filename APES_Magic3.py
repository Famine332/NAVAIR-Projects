# © 2020 Robert Leonard, All Rights Reserved
from openpyxl import load_workbook
from collections import OrderedDict
from openpyxl.styles import NamedStyle
import openpyxl
import os
import re

# Find the active directory in use
path = os.path.dirname(os.path.abspath(__file__))

# list the files in that directory
files = os.listdir(path)

# Find the file that contains 'HQMC' and append to list
matching = [s for s in files if "HQMC" in s]

# Assign variable HQMC to first match, second match is likely the ghost file if it has already been opened in excel
HQMC = matching[0]

# Concatenate active path with HQMC file to get full path of the file
path = path+"\\"
HQMC_path = path+HQMC
wb = load_workbook(HQMC_path)

# Set eh active worksheet to HQMC file
ws = wb.active
# define the following variables as lists so I can append values to them
data = []
data_rows = []
# establish a range of row values to determine what row the BUNOs start
for i in range(15, 30):
    if 'BUNO' in str(ws.cell(row=i, column=4).value):
        # Iterate rows under BUNO to collect the BUNO numbers and the row number
        for x in range(i+1, 50):
            data.append(ws.cell(row=x, column=4).value)
            # data_rows captures the excel row the data resides
            data_rows.append(x)
# Build a dictionary to pair BuNo entry with EXCEL row number in the HQMC report
HQMC_dict = dict(zip(data, data_rows))
# Delete all rows that are not BuNos
del HQMC_dict[None]

# Sequence of filters to remove python file, PMA file, and any tilda file from list, respectively


def remove_pyfile(filename):
    return '.py' not in filename


list1 = (list(filter(lambda x: remove_pyfile(x), files)))


def remove_PMAfile(filename):
    return 'PMA' not in filename


list2 = (list(filter(lambda x: remove_PMAfile(x), list1)))


def remove_tilda_file(filename):
    return '~' not in filename


file_list = (list(filter(lambda x: remove_tilda_file(x), list2)))

# Define the below variables as lists to I can later append values
thebunos = []
thecomments = []
eecd = []
recd = []
current = []

# Iterate through each file in the directory where the python script resides
for file in file_list:
    # Load the workbook
    wb = load_workbook(path+file, data_only=True)
    # Iterate through each sheet in the workbook
    for ws in wb.worksheets:
        # check each sheet to see if cell D2 is 'BUNO' / this is how we known
        # which sheets to collect data from, because those sheets have the
        # BUNO title in that cell
        if ws.cell(row=2, column=4).value == "BUNO":
            # Iterate through all populated rows on that worksheet
            for x in range(1, ws.max_row+1):
                # Append the BUNO number to the buno list
                thebunos.append(ws.cell(row=x, column=4).value)
                # Set the range of columns to check for other data
                for i in range(5, 26):
                    # Find the column for comments
                    if 'Comments'in str(ws.cell(row=2, column=i).value):
                        # Append the comment in that column to the list
                        thecomments.append(ws.cell(row=x, column=i).value)
                    # Find the column for EECD
                    if 'EECD' == str(ws.cell(row=2, column=i).value):
                        # Append the EECD date to the list
                        eecd.append(ws.cell(row=x, column=i).value)
                    # Find the column for the RECD
                    if 'RECD' == str(ws.cell(row=2, column=i).value):
                        # Append the RECD date to the list
                        recd.append(ws.cell(row=x, column=i).value)
                    # Find the column for the Customer Expected delivery date
                    if 'Current Date Customer'in str(ws.cell(row=2, column=i).value):
                        # Append that date to the list
                        current.append(ws.cell(row=x, column=i).value)
# Create dictionaries from the lists, with the key = to the BUNO and the value
# equal to the respective data or comment
COMMENTS = dict(zip(thebunos, thecomments))
EECD = dict(zip(thebunos, eecd))
RECD = dict(zip(thebunos, recd))
CURRENT = dict(zip(thebunos, current))


# Load the HQMC workbook for editing
wb = load_workbook(HQMC_path)
ws = wb.active
# Iteration of 'i' will search the B-column for the row where the title 'Site'
# starts in order to align the below titles to that same line
for i in range(15, 30):
    if 'Site'in str(ws.cell(row=i, column=2).value):
        ws['P'+str(i)] = 'EECD'
        ws['Q'+str(i)] = 'RECD'
        ws['R'+str(i)] = 'Customer Expects'
        ws['S'+str(i)] = 'Comments'

# Loop through the dictionaries to write EECD/RECD/Customer Expected/Comments
for key in HQMC_dict.keys():
    rownum = HQMC_dict[key]
    try:
        ws['P'+str(rownum)] = EECD[key]
    except:
        pass
    try:
        ws['Q'+str(rownum)] = RECD[key]
    except:
        pass
    try:
        ws['R'+str(rownum)] = CURRENT[key]
    except:
        pass
    try:
        ws['S'+str(rownum)] = COMMENTS[key]
    except:
        pass
wb.save(path+'PMA-265 HQMC Update ().xlsx')
